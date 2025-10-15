import os
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from utils.conciliacao import conciliar_movimentos_db  # ✅ sem import circular
from models.user_model import MovimentacaoBAI, MovimentacaoContabilidade, ExecucaoReconciliacao

def gerar_pdf_conciliacao(db, execucao_id: int, caminho_pdf: str):
    # 🔹 Garante que a pasta existe
    pasta = os.path.dirname(caminho_pdf)
    os.makedirs(pasta, exist_ok=True)

    doc = SimpleDocTemplate(caminho_pdf, pagesize=A4)
    elementos = []
    styles = getSampleStyleSheet()

    # 🔹 Buscar dados no banco
    execucao = db.query(ExecucaoReconciliacao).filter_by(id=execucao_id).first()
    extrato = db.query(MovimentacaoBAI).filter_by(execucao_id=execucao_id).all()
    contab = db.query(MovimentacaoContabilidade).filter_by(execucao_id=execucao_id).all()

    # 🔹 Cabeçalho
    elementos.append(Paragraph("Relatório de Conciliação", styles['Title']))
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph(f"Empresa ID: {execucao.empresa_id}", styles['Normal']))
    elementos.append(Paragraph(f"Execução ID: {execucao.id}", styles['Normal']))
    elementos.append(Paragraph(f"Data: {execucao.criado_em}", styles['Normal']))
    elementos.append(Spacer(1, 24))

    # 🔹 Resumo
    data_resumo = [
        ["Total Extrato", len(extrato)],
        ["Total Contabilidade", len(contab)],
    ]
    tabela_resumo = Table(data_resumo, colWidths=[200, 200])
    tabela_resumo.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ])
    )
    elementos.append(tabela_resumo)
    elementos.append(Spacer(1, 24))

    # 🔹 Tabela simplificada (extrato)
    data_tabela = [["Data", "Descrição", "Débito", "Crédito", "Saldo"]]
    for mov in extrato[:10]:  # ← limitar só p/ exemplo
        data_tabela.append([
            str(mov.data_mov),
            mov.descritivo or "",
            mov.debito or 0,
            mov.credito or 0,
            mov.saldo or 0
        ])

    tabela = Table(data_tabela, colWidths=[80, 200, 80, 80, 80])
    tabela.setStyle(
        TableStyle([
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ])
    )
    elementos.append(tabela)

    doc.build(elementos)


def gerar_excel_conciliacao(db, execucao_id: int, caminho_excel: str):
    # 🔹 Garante que a pasta existe
    pasta = os.path.dirname(caminho_excel)
    os.makedirs(pasta, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    # 🔹 Buscar execução
    execucao = db.query(ExecucaoReconciliacao).filter_by(id=execucao_id).first()

    # 🔹 Rodar conciliação (usando dados da execução)
    conciliacao = conciliar_movimentos_db(db, execucao_id)

    # 🔹 Filtro de palavras a ignorar (igual ao processar_contabil)
    palavras_ignoradas = {"SALDO INICIAL", "SALDO FINAL", "TRANSPORTE", "A TRANSPORTAR"}

    def linha_valida(descritivo: str) -> bool:
        if not descritivo:
            return True
        return not any(p in descritivo.upper() for p in palavras_ignoradas)

    # --- CABEÇALHO ---
    ws.append([f"Relatório de Conciliação - Execução {execucao.id}"])
    ws.append([f"Empresa ID: {execucao.empresa_id}", f"Data: {execucao.criado_em}"])
    ws.append([])

    # --- RESUMO ---
    resumo = conciliacao["summary"]
    ws.append(["Categoria", "Quantidade"])
    ws.append(["Total Extrato", resumo["total_extrato"]])
    ws.append(["Total Contabilidade", resumo["total_contabilidade"]])
    ws.append(["Conciliados", resumo["conciliados"]])
    ws.append(["Somente Extrato", resumo["somente_extrato"]])
    ws.append(["Somente Contabilidade", resumo["somente_contabilidade"]])
    ws.append([])

    # --- DETALHES ---
    ws.append(["Origem", "Data Mov", "Descrição", "Débito", "Crédito", "Valor Líquido", "Status"])

    # Cores para diferenciar
    fill_conciliado = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde
    fill_divergente = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # amarelo
    fill_extrato = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")     # azul claro
    fill_contab = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")     # vermelho claro

    # Conciliados / divergentes
    for c in conciliacao["conciliados"]:
        if not linha_valida(c["extrato_descritivo"]) and not linha_valida(c["contab_descritivo"]):
            continue  # ignora linha auxiliar

        row = [
            "Conciliado",
            c["extrato_data_mov"] or c["contab_data_mov"],
            f"{c['extrato_descritivo']} | {c['contab_descritivo']}",
            c["extrato_valor_liq"],
            c["contab_valor_liq"],
            c["extrato_valor_liq"] - c["contab_valor_liq"],
            c["status"],
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.fill = fill_conciliado if c["status"] == "conciliado" else fill_divergente

    # Somente extrato
    for e in conciliacao["somente_extrato"]:
        if not linha_valida(e["descritivo"]):
            continue

        row = [
            "Somente Extrato",
            e["data_mov"],
            e["descritivo"],
            e["debito"],
            e["credito"],
            e["valor_liq"],
            "Não conciliado",
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.fill = fill_extrato

    # Somente contabilidade
    for c in conciliacao["somente_contabilidade"]:
        if not linha_valida(c["descritivo"]):
            continue

        row = [
            "Somente Contabilidade",
            c["data_mov"],
            c["descritivo"],
            c["debito"],
            c["credito"],
            c["valor_liq"],
            "Não conciliado",
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.fill = fill_contab

    wb.save(caminho_excel)